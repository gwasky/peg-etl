from openpyxl import load_workbook
import os
import sys
import MySQLdb
import logging
import datetime

# CREATE TABLE `peg`.`customers`  (
#   `customerId` int(11) NOT NULL,
#   `location` varchar(50) NOT NULL,
#   `first_name` varchar(50) NOT NULL,
#   `last_name` varchar(50) NOT NULL,
#   `end_date` datetime(0) NULL DEFAULT NULL,
#   PRIMARY KEY (`customerId`) USING BTREE,
#   INDEX `customerId_idx`(`customerId`) USING BTREE,
#   INDEX `end_date_idx`(`end_date`) USING BTREE
# ) ENGINE = InnoDB CHARACTER SET = latin1 COLLATE = latin1_swedish_ci ROW_FORMAT = Dynamic;

## Prepare statements would be more optimal, in the interest of time i opted not to use them

logger = logging.getLogger("PEG_ETL")
logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s;%(levelname)s;%(message)s")
ch.setFormatter(formatter)
logger.addHandler(ch)

def customerExists(id):
    logger.info(str(id) + ' Checking existing data for Active Record')
    cursor = db.cursor()
    cursor.execute("SELECT * FROM customers WHERE customerId = "+ str(id) +" AND end_date IS NULL")
    result_set = cursor.fetchall()
    for row in result_set:
        logger.info(" %s Found | - Current Records are : %s |  %s | %s | %s " % (row[0], row[1], row[2], row[3], row[4]))
        cursor.close()
    return result_set


def deactivateCurrentRecord(id):
    result = False
    cursor = db.cursor()
    try:
        cursor.execute("""UPDATE customers SET end_date = %s WHERE customerId = %s """ , ( datetime.datetime.now() , id ))
        db.commit()
        result = True
        cursor.close()
    except (MySQLdb.Error, MySQLdb.Warning) as ex:
        logger.error(ex)
        return None
    return result

def insertNewRecord(CustomerId, Location, First_Name, Last_Name):
    result = False
    cursor = db.cursor()
    try:
        cursor.execute("""INSERT INTO customers (customerId,location,first_name,last_name) VALUES ( %s , '%s' , '%s' , '%s') """ % \
        (CustomerId, Location, First_Name, Last_Name))
        db.commit()
        result = True
        cursor.close()
    except (MySQLdb.Error, MySQLdb.Warning) as ex:
        logger.error(ex)
        cursor.close()
    return result

os.getcwd()
os.chdir("D:\personal\PEGAfrica\Senior DE Engineer")
wb = load_workbook('customer-changing-data.xlsx')
ws = wb['Sheet1']

db = MySQLdb.connect("localhost","root","", "peg")

logger.info(" Successfully Connected to the Database")
# print)

logger.info(" Fetching New Excel Data from Row 18 to 27")

for row in range(18, 27):
    CustomerId = ws['A' + str(row)].value
    Location = ws['B' + str(row)].value
    First_Name = ws['C' + str(row)].value
    Last_Name = ws['D' + str(row)].value

    result_set = customerExists(CustomerId)
    if(result_set):
        for row in result_set:
            logger.info(str(row[0]) + " Checking if new records are different from the old records for customer with Id")
            if row[1] != Location or row[2] != First_Name or row[3] != Last_Name:
                logger.info(str(row[0]) + " None Existing Entry Found")
                logger.info(str(row[0]) + " Insert Response | " + str(deactivateCurrentRecord(row[0])))
                result = insertNewRecord(CustomerId, Location, First_Name, Last_Name)
                logger.info(str(row[0]) + " Insert Response | " + str(result))
            else:
                logger.info(str(row[0]) + " Duplicate Entry")
    else:
        logger.info(str(row[0]) + " New Record")
        result = insertNewRecord(CustomerId, Location, First_Name, Last_Name)
        logger.info(str(row[0]) + " New Record Insert Response | " + str(result))

db.close()
    