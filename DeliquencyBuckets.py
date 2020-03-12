from openpyxl import load_workbook
import os
import string
import time
import sys
import datetime

def get_no_light_band(days_without_light): 
    band = ''
    if days_without_light >= 0 and days_without_light <= 31:
        band = '0-31'
    elif days_without_light >= 32 and days_without_light <= 62:
        band = '32-62'
    elif days_without_light > 62 and days_without_light <= 93:
        band = '63-93'
    elif days_without_light > 93 and days_without_light <= 124:
        band = '94-124'
    elif days_without_light > 124 and days_without_light <= 155:
        band = '125-155'
    else:
        band = '156+'
    return band

def get_provisioning_percentage(band, status = ''):
    provisioning_percentage = 0
    switcher = {
        "156+" : 100,
        "125-155" : 85,
        "0-31" : 0,
        "32-62" : 50,
        "94-124" : 65,
        "63-93" : 50,
    }
    provisioning_percentage = switcher.get(band)
    if status == 'Cancelled':
        provisioning_percentage = 100
    elif status == 'Finished Payment':
        provisioning_percentage = 0
    return provisioning_percentage

os.getcwd()
os.chdir("D:\personal\PEGAfrica\Senior DE Engineer")
wb = load_workbook('provisioning-data-1.xlsx')
ws = wb['DATA']

data_struct = {}

for row in range(2, ws.max_row):
    Loan_Id = ws['A' + str(row)].value
    Date_Of_Activity = ws['B' + str(row)].value
    Credits_End_Date = ws['C' + str(row)].value
    Outstanding_Balance = ws['D' + str(row)].value
    customerStatus = ws['E' + str(row)].value
    parProductType = ws['F' + str(row)].value

    # Remove Null
    Outstanding_Balance = Outstanding_Balance or 0

    Days_Without_Light = abs((Date_Of_Activity - Credits_End_Date).days)
    No_light_band = get_no_light_band(abs((Date_Of_Activity - Credits_End_Date).days))

    ProvisioningPercentage = get_provisioning_percentage(
        get_no_light_band(abs((Date_Of_Activity - Credits_End_Date).days)),customerStatus)

    Provisioning_Amount = Outstanding_Balance * (ProvisioningPercentage/100)

    data_struct.setdefault(Date_Of_Activity, {})
    data_struct[Date_Of_Activity].setdefault(Loan_Id, {})
    data_struct[Date_Of_Activity][Loan_Id]['Credits_End_Date'] = Credits_End_Date
    data_struct[Date_Of_Activity][Loan_Id]['Outstanding_Balance'] = Outstanding_Balance
    data_struct[Date_Of_Activity][Loan_Id]['customerStatus'] = customerStatus
    data_struct[Date_Of_Activity][Loan_Id]['ProductType'] = parProductType
    data_struct[Date_Of_Activity][Loan_Id]['Days without Light'] = Days_Without_Light
    data_struct[Date_Of_Activity][Loan_Id]['Band'] = No_light_band
    data_struct[Date_Of_Activity][Loan_Id]['Provisioning Percentage'] = ProvisioningPercentage
    data_struct[Date_Of_Activity][Loan_Id]['Provisioning Amount'] = Provisioning_Amount

print(data_struct[datetime.datetime(2019, 6, 30, 0, 0)][1250598])

prov_struct = {}
for ActivityDate in data_struct:
    print(ActivityDate)
    if ActivityDate == datetime.datetime(2019, 6, 30, 0, 0):
        print('----- COMPUTING OPENING PROVISION AMOUNT-----')
        for Loan_Id in data_struct[ActivityDate]:

            prov_struct.setdefault('Opening', {}). \
            setdefault(data_struct[ActivityDate][Loan_Id]['ProductType'],{}). \
            setdefault(data_struct[ActivityDate][Loan_Id]['Band'], {})
            if data_struct[ActivityDate][Loan_Id]['Days without Light'] > 0:
                prov_struct['Opening'][data_struct[ActivityDate][Loan_Id]['ProductType']][data_struct[ActivityDate][Loan_Id]['Band']] = + \
                data_struct[ActivityDate][Loan_Id]['Provisioning Amount']

    if ActivityDate == datetime.datetime(2019, 7, 22, 0, 0):
        print('----- COMPUTING CLOSING PROVISION AMOUNT-----')
        for Loan_Id in data_struct[ActivityDate]:
            prov_struct.setdefault('Closing', {}). \
            setdefault(data_struct[ActivityDate][Loan_Id]['ProductType'],{}). \
            setdefault(data_struct[ActivityDate][Loan_Id]['Band'],{})
            prov_struct['Closing'][data_struct[ActivityDate][Loan_Id]['ProductType']][data_struct[ActivityDate][Loan_Id]['Band']] =+ \
                data_struct[ActivityDate][Loan_Id]['Provisioning Amount']


print(prov_struct)